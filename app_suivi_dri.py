import streamlit as st
import pandas as pd
import geopandas as gpd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime
import shutil
import zipfile
import os
import tempfile
from io import BytesIO
import requests
import base64
import time
from openpyxl.styles import PatternFill, Font
import plotly.express as px
import fiona
import zipfile
from fastkml import kml

def traiter_gdb_thd_zones(fichier_gdb_zip):
    import tempfile
    import zipfile
    import os

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            zip_path = os.path.join(tmpdir, "gdb.zip")
            with open(zip_path, "wb") as f:
                f.write(fichier_gdb_zip.read())

            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(tmpdir)

            # Trouver le dossier .gdb
            gdb_dirs = [os.path.join(tmpdir, d) for d in os.listdir(tmpdir) if d.endswith(".gdb")]
            if not gdb_dirs:
                return None, "Aucun dossier .gdb trouvé dans l'archive"
            gdb_path = gdb_dirs[0]

            import fiona
            import geopandas as gpd

            # Identifier une couche contenant les colonnes nécessaires
            layers = fiona.listlayers(gdb_path)
            selected_layer = None
            for layer in layers:
                sample = gpd.read_file(gdb_path, layer=layer).head(1)
                if all(col in sample.columns for col in ["DSP", "ID_DSP", "Z_BPE"]):
                    selected_layer = layer
                    break

            if not selected_layer:
                return None, "Aucune couche avec colonnes DSP, ID_DSP, Z_BPE trouvée"

            gdf = gpd.read_file(gdb_path, layer=selected_layer)
            gdf = gdf.to_crs(epsg=2154)

            zones_valides = [f"THD ZONE {i}" for i in range(1, 5)]
            gdf = gdf[gdf["Z_BPE"].isin(zones_valides)].copy()
            gdf["geometry"] = gdf.buffer(5)

            zones_bufferisees = {
                zone: gdf[gdf["Z_BPE"] == zone].copy()
                for zone in zones_valides
            }

            return zones_bufferisees, None

    except Exception as e:
        return None, f"Erreur lors du traitement du fichier GDB : {e}"

def creer_zip_final(chemin_shp1, chemin_shp2, image_tcd_buf, nom_commande):
    import shutil
    import glob

    try:
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, f"{nom_commande}_EXPORT.zip")

        # Récupérer tous les fichiers liés aux SHP (shp, shx, dbf, prj, cpg, etc.)
        def copier_shp(chemin_shp, dossier_dest, prefix):
            base = chemin_shp.replace(".shp", "")
            for f in glob.glob(base + ".*"):
                ext = os.path.splitext(f)[1]
                dest = os.path.join(dossier_dest, f"{prefix}{ext}")
                shutil.copy(f, dest)

        copier_shp(chemin_shp1, temp_dir, f"BPE_{nom_commande}")
        copier_shp(chemin_shp2, temp_dir, f"SYNTH_{nom_commande}")

        # Enregistrer image PNG
        image_path = os.path.join(temp_dir, f"TCD_MA_{nom_commande}.png")
        with open(image_path, "wb") as f:
            f.write(image_tcd_buf.getbuffer())

        # Création du ZIP
        with zipfile.ZipFile(zip_path, "w") as zipf:
            for filename in os.listdir(temp_dir):
                if filename != os.path.basename(zip_path):  # ne pas inclure le zip lui-même
                    zipf.write(os.path.join(temp_dir, filename), arcname=filename)

        return zip_path, None

    except Exception as e:
        return None, f"Erreur création ZIP : {e}"


def generer_image_tcd_ma(gdf_synthese, nom_commande):
    import matplotlib.pyplot as plt
    from io import BytesIO

    try:
        # TCD
        pivot = pd.pivot_table(
            gdf_synthese,
            index="NB_SALARIE",
            columns="GAIN_ENTREPRISE",
            values="SIRET",
            aggfunc="count",
            margins=True,
            margins_name="Total général",
            fill_value=0
        )

        ordered_cols = [
            "Nouvellement forfaitaire",
            "Changement forfaitaire",
            "Pas de changement"
        ]
        existing_cols = [col for col in ordered_cols if col in pivot.columns]
        pivot = pivot[existing_cols + ["Total général"]]
        pivot = pivot.sort_index()

        # Résumé
        counts = gdf_synthese["GAIN_ENTREPRISE"].value_counts()
        phrases = []
        if "Nouvellement forfaitaire" in counts:
            phrases.append(f"- {counts['Nouvellement forfaitaire']} entreprises deviennent éligibles forfaitairement")
        if "Changement forfaitaire" in counts:
            phrases.append(f"- {counts['Changement forfaitaire']} entreprises changent de zone forfaitaire positivement")

        # Image matplotlib
        fig, ax = plt.subplots(figsize=(11, len(pivot) * 0.6 + len(phrases) * 0.6 + 2))
        y_offset = 0.1
        for i, phrase in enumerate(phrases):
            ax.text(0, 1 - y_offset * (i + 1), phrase, fontsize=10, ha='left')
        ax.axis('off')

        table_data = [[idx] + list(row) for idx, row in pivot.iterrows()]
        col_labels = ["NB_SALARIE"] + pivot.columns.to_list()

        table = ax.table(
            cellText=table_data,
            colLabels=col_labels,
            loc='center',
            cellLoc='center'
        )
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.2, 1.3)

        for key, cell in table.get_celld().items():
            row, col = key
            cell.set_text_props(wrap=True)
            if row == 0:
                cell.set_facecolor('#d3d3d3')
                cell.set_text_props(weight='bold')
            elif col == 0:
                cell.set_facecolor('#f0f0f0')
            if row > 0 and pivot.index[row - 1] == "Total général":
                cell.set_facecolor('#f5e6cc')
                cell.set_text_props(weight='bold')

        # Enregistrer dans buffer mémoire
        buf = BytesIO()
        plt.savefig(buf, format="png", bbox_inches='tight', dpi=300)
        buf.seek(0)
        return buf, None

    except Exception as e:
        return None, f"Erreur génération image TCD : {e}"

def generer_synthese_finale(gdf, nom_commande):
    try:
        gdf_synthese = gdf.copy()

        # Nettoyage colonnes
        cols_to_drop = [col for col in gdf_synthese.columns if "_predicate" in col or "l1_normali" in col]
        gdf_synthese = gdf_synthese.drop(columns=cols_to_drop, errors="ignore")

        # Arrondir DISTANCE + conversion
        gdf_synthese["DISTANCE"] = gdf_synthese["DISTANCE"].round().astype(int)

        # Renommer Z_BPE
        gdf_synthese = gdf_synthese.rename(columns={"Z_BPE": "THD_ACTUELLE"})

        # Générer THD_FINALE
        def get_thd_finale(actuelle, extension):
            if not actuelle or not extension:
                return "INDETERMINE"
            try:
                num_act = int(actuelle[-1])
                num_ext = int(extension[-1])
                if num_ext <= num_act:
                    return actuelle
                else:
                    return f"THD {actuelle[-1]} A {extension[-1]}"
            except:
                return "INDETERMINE"
        gdf_synthese["THD_FINALE"] = gdf_synthese.apply(lambda row: get_thd_finale(row["THD_ACTUELLE"], row["THD_EXTENSION"]), axis=1)

        # GAIN_ENTREPRISE
        def get_gain(val):
            if "5 A" in val:
                return "Nouvellement forfaitaire"
            elif "ZONE" in val:
                return "Pas de changement"
            else:
                return "Changement forfaitaire"
        gdf_synthese["GAIN_ENTREPRISE"] = gdf_synthese["THD_FINALE"].apply(get_gain)

        # Ajouter commande
        gdf_synthese["COMMANDE"] = nom_commande

        # Export SHP
        export_dir = tempfile.mkdtemp()
        shp_path = os.path.join(export_dir, f"{nom_commande}_synthese.shp")
        gdf_synthese.to_file(shp_path, driver="ESRI Shapefile")

        return gdf_synthese, shp_path, None

    except Exception as e:
        return None, None, f"Erreur synthèse finale : {e}"

def attribuer_thd_extension(gdf_bpe, thd_buffers):
    try:
        gdf_result = gdf_bpe.copy()

        # Distance minimale à un BPE
        bpe_points = gdf_bpe.geometry
        gdf_result["DISTANCE"] = gdf_result.geometry.apply(lambda geom: bpe_points.distance(geom).min())
        gdf_result["DISTANCE"] = gdf_result["DISTANCE"].round().astype(int)

        # THD_EXTENSION par seuils
        def get_thd_zone(dist):
            if dist < 100:
                return "THD ZONE 1"
            elif dist < 200:
                return "THD ZONE 2"
            elif dist < 300:
                return "THD ZONE 3"
            elif dist < 500:
                return "THD ZONE 4"
            else:
                return None
        gdf_result["THD_EXTENSION"] = gdf_result["DISTANCE"].apply(get_thd_zone)

        # THD actuelle (Z_BPE) via croisement spatial avec buffers
        gdf_result["Z_BPE"] = None
        for zone_name, buffer_gdf in thd_buffers.items():
            matched = gpd.sjoin(gdf_result[gdf_result["Z_BPE"].isna()], buffer_gdf, how="inner", predicate="intersects")
            if not matched.empty:
                gdf_result.loc[matched.index, "Z_BPE"] = zone_name

        # Ce qui reste non assigné = Z_BPE = THD ZONE 5
        gdf_result["Z_BPE"] = gdf_result["Z_BPE"].fillna("THD ZONE 5")

        return gdf_result, None

    except Exception as e:
        return None, f"Erreur attribution THD : {e}"


def traiter_bpe_shp(fichier_shp, nom_commande):
    try:
        gdf = gpd.read_file(fichier_shp)

        # Vérifie géométrie POINT
        if gdf.geom_type.unique()[0] != "Point":
            return None, "Le fichier BPE doit contenir uniquement des points."

        # Garde seulement la colonne utile
        if "Name" in gdf.columns:
            gdf = gdf[["Name", "geometry"]].rename(columns={"Name": "BPE"})
        elif "BPE" in gdf.columns:
            gdf = gdf[["BPE", "geometry"]]
        else:
            return None, "Colonne 'Name' ou 'BPE' absente du shapefile."

        # Ajoute la colonne COMMANDE
        gdf["COMMANDE"] = nom_commande

        # Projection en 2154
        gdf = gdf.to_crs(epsg=2154)

        # Export SHP
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, f"BPE_{nom_commande}.shp")
        gdf.to_file(output_path)

        return gdf, output_path

    except Exception as e:
        return None, f"Erreur lors du traitement du fichier SHP : {e}"

def traiter_gdb_thd_zones(fichier_gdb):
    import tempfile

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            # Sauvegarder le fichier GDB temporairement
            gdb_path = os.path.join(tmpdir, "fichier.gdb")
            if fichier_gdb.name.endswith(".zip"):
                with zipfile.ZipFile(fichier_gdb, "r") as zip_ref:
                    zip_ref.extractall(gdb_path)
            else:
                with zipfile.ZipFile(fichier_gdb, "r") as zip_ref:
                    zip_ref.extractall(tmpdir)
                
                # Localise le dossier .gdb
                gdb_dirs = [os.path.join(tmpdir, d) for d in os.listdir(tmpdir) if d.endswith(".gdb")]
                if not gdb_dirs:
                    return None, "Aucun dossier .gdb trouvé dans l'archive"
                gdb_path = gdb_dirs[0]

            # Chercher une couche avec les colonnes nécessaires
            layers = fiona.listlayers(gdb_path)
            selected_layer = None
            for layer in layers:
                sample = gpd.read_file(gdb_path, layer=layer).head(1)
                if all(col in sample.columns for col in ["DSP", "ID_DSP", "Z_BPE"]):
                    selected_layer = layer
                    break

            if not selected_layer:
                return None, "Aucune couche avec colonnes DSP, ID_DSP, Z_BPE trouvée."

            gdf = gpd.read_file(gdb_path, layer=selected_layer)
            gdf = gdf.to_crs(epsg=2154)

            # On filtre pour les Z_BPE valides
            zones_valides = [f"THD ZONE {i}" for i in range(1, 5)]
            gdf = gdf[gdf["Z_BPE"].isin(zones_valides)].copy()

            # Buffer 5m
            gdf["geometry"] = gdf.buffer(5)

            # Dictionnaire par zone
            zones_bufferisees = {
                zone: gdf[gdf["Z_BPE"] == zone].copy()
                for zone in zones_valides
            }

            return zones_bufferisees, None

    except Exception as e:
        return None, f"Erreur lors du traitement GDB : {e}"

# --- Fonction pour push sur GitHub ---
def push_to_github(token, repo, path_in_repo, local_filepath, commit_message="Mise à jour fichier DRI"):
    api_url = f"https://api.github.com/repos/{repo}/contents/{path_in_repo}"
    headers = {"Authorization": f"token {token}"}

    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        sha = response.json()["sha"]
    else:
        sha = None

    with open(local_filepath, "rb") as f:
        content = f.read()
        encoded_content = base64.b64encode(content).decode("utf-8")

    data = {
        "message": commit_message,
        "content": encoded_content,
        "branch": "main"
    }

    if sha:
        data["sha"] = sha

    put_response = requests.put(api_url, headers=headers, json=data)
    return put_response.status_code, put_response.json()

# --- Fonction d'appel simplifiée ---
def try_push_to_github():
    try:
        token = st.secrets["GITHUB_TOKEN"]
        repo = "Ali-geomarket/TRAITEMENT-FICHE-DRI"
        path_in_repo = "Suivi_demandes_AUTOMATISATION.xlsx"
        status, result = push_to_github(token, repo, path_in_repo, suivi_file_path)
        if status in [200, 201]:
            st.success("Fichier synchronisé avec GitHub")
        else:
            st.warning(f"Push GitHub échoué : {result}")
    except Exception as e:
        st.warning(f"Erreur lors du push GitHub : {e}")

# --- Authentification ---
USERS = {"sg": "dri", "ps": "dri"}

st.set_page_config(page_title="Fiche DRI & TCD MA", layout="wide")

# --- Session State ---
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
if "main_page" not in st.session_state:
    st.session_state["main_page"] = "login"
if "ligne_temporaire" not in st.session_state:
    st.session_state["ligne_temporaire"] = None

# --- Constantes ---
COLONNES = [
    "DATE RECEPTION", "RESEAU", "RESPONSABLE PROD", "COMMERCIAL", "PROJET", "TYPE DE DEMANDE",
    "COUT EXTENSION", "COUT GLOBAL PROJET", "OPERATEUR", "GAIN DRI", "ROI", "NB CLIENTS AMORTISSEMENT",
    "COMMANDE", "DATE TRAITEMENT", "DELAI TRAITEMENT", "ETAT GEOMARKETING", "RESP GEOMARKET",
    "CONCLUSION", "COMMENTAIRE"
]

suivi_file_path = "Suivi_demandes_AUTOMATISATION.xlsx"
temp_export_path = "Suivi_demandes_EXPORT.xlsx"

# -------------------------
# PAGE : CONNEXION
# -------------------------
if not st.session_state["authenticated"]:
    st.title("Connexion requise")
    with st.form("login_form"):
        username = st.text_input("Nom d'utilisateur")
        password = st.text_input("Mot de passe", type="password")
        if st.form_submit_button("Se connecter"):
            if username in USERS and USERS[username] == password:
                st.session_state["authenticated"] = True
                st.session_state["user"] = username
                st.session_state["main_page"] = "home"
                st.session_state["current_section"] = "visualisation_modif" 
                st.rerun()
            else:
                st.error("Identifiants incorrects")

# -------------------------
# INTERFACE PRINCIPALE APRÈS CONNEXION
# -------------------------
elif st.session_state["authenticated"]:
    st.title("Application Fiche DRI & TCD MA")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button("Visualisation globale & Modification"):
            st.session_state["current_section"] = "visualisation_modif"
    with col2:
        if st.button("Ajouter une nouvelle fiche DRI"):
            st.session_state["current_section"] = "ajout"
            st.session_state["ligne_temporaire"] = None
    
    with col4:
        if st.button("Analyse des données"):
            st.session_state["current_section"] = "analyse_donnees"

    # Ensuite on affiche la section choisie :
    section = st.session_state.get("current_section")

    if section == "visualisation_modif":
        st.subheader("Visualisation & Modification du fichier Excel")
    
        if "show_full_df" not in st.session_state:
            st.session_state["show_full_df"] = False
    
        try:
            df = pd.read_excel(suivi_file_path, engine="openpyxl")
            df.columns = [col.strip().upper() for col in df.columns]
            df = df.reindex(columns=COLONNES)
    
            df["DATE RECEPTION"] = pd.to_datetime(df["DATE RECEPTION"], errors="coerce", dayfirst=True)
            df["DATE TRAITEMENT"] = pd.to_datetime(df["DATE TRAITEMENT"], errors="coerce", dayfirst=True)
    
            toggle_label = "Afficher tout le fichier" if not st.session_state["show_full_df"] else "Afficher un extrait"
            toggle_clicked = st.button(toggle_label)
            if toggle_clicked:
                st.session_state["show_full_df"] = not st.session_state["show_full_df"]
                st.rerun()
    
            if st.session_state["show_full_df"]:
                display_df = df.copy()
            else:
                display_df = df.tail(15).copy()
    
            edited_df = st.data_editor(display_df, use_container_width=True, num_rows="dynamic")
    
            if not st.session_state["show_full_df"]:
                df_update = df.copy()
                df_tail_idx = df.tail(15).index
                for i, idx in enumerate(df_tail_idx):
                    df_update.loc[idx] = edited_df.iloc[i]
            else:
                df_update = edited_df.copy()
    
            if st.button("Enregistrer les modifications"):
                try:
                    full_df = pd.read_excel(suivi_file_path, engine="openpyxl")
                    full_df.columns = [col.strip().upper() for col in full_df.columns]
                    full_df = full_df.reindex(columns=COLONNES)
    
                    if not st.session_state["show_full_df"]:
                        tail_indices = full_df.tail(15).index
                        for i, idx in enumerate(tail_indices):
                            full_df.loc[idx] = edited_df.iloc[i]
                    else:
                        full_df = edited_df.copy()
    
                    wb = load_workbook(suivi_file_path)
                    ws = wb.active
                    ws.delete_rows(2, ws.max_row)
    
                    for _, row in full_df.iterrows():
                        cleaned_row = [None if pd.isna(cell) else cell for cell in row]
                        ws.append(cleaned_row)
    
                    headers = [cell.value for cell in ws[1]]
                    if "ETAT GEOMARKETING" in headers:
                        etat_col_index = headers.index("ETAT GEOMARKETING")
                        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                            cell = row[etat_col_index]
                            if cell.value == "OK":
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell.font = Font(color="006100")
    
                    wb.save(suivi_file_path)
                    try_push_to_github()
                    st.success("Fichier mis à jour avec succès")
    
                except Exception as e:
                    st.error(f"Erreur lors de la sauvegarde : {e}")
    
            shutil.copy(suivi_file_path, temp_export_path)
            st.download_button("Télécharger le fichier Excel", data=open(temp_export_path, "rb"),
                               file_name="Suivi_demandes_EXPORT.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with st.expander("Ajouter une date de traitement"):
                commande_input = st.text_input("Entrez la commande (ex: CMD_X_00003320)")
                date_traitement_input = st.date_input("Date de traitement", format="DD/MM/YYYY")
            
                if st.button("Mettre à jour la date de traitement"):
                    try:
                        df = pd.read_excel(suivi_file_path, engine="openpyxl")
                        df.columns = [col.strip().upper() for col in df.columns]
            
                        occurrences = df[df["COMMANDE"] == commande_input]
            
                        if len(occurrences) == 0:
                            st.error("Commande introuvable.")
                        elif len(occurrences) > 1:
                            st.error(f"Il y a {len(occurrences)} lignes avec la commande '{commande_input}'. Mise à jour annulée.")
                        else:
                            # Une seule ligne → on poursuit
                            idx = occurrences.index[0]
                            df.at[idx, "DATE TRAITEMENT"] = pd.to_datetime(date_traitement_input)
            
                            dt_traitement = pd.to_datetime(df.at[idx, "DATE TRAITEMENT"], errors="coerce")
                            dt_reception = pd.to_datetime(df.at[idx, "DATE RECEPTION"], errors="coerce")
            
                            if pd.notnull(dt_traitement) and pd.notnull(dt_reception):
                                delai = (dt_traitement - dt_reception).days
                                df.at[idx, "DELAI TRAITEMENT"] = delai
                                st.success(f"Date et délai mis à jour : {delai} jours")
            
                                # Mise à jour dans le fichier Excel
                                wb = load_workbook(suivi_file_path)
                                ws = wb.active
                                headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
                                for row in ws.iter_rows(min_row=2):
                                    if row[headers.index("COMMANDE")].value == commande_input:
                                        row[headers.index("DATE TRAITEMENT")].value = date_traitement_input.strftime("%d/%m/%Y")
                                        row[headers.index("DELAI TRAITEMENT")].value = delai
                                        break
            
                                wb.save(suivi_file_path)
                                try_push_to_github()
                                st.rerun()
                            else:
                                st.warning("Dates invalides : réception ou traitement manquante.")
                    except Exception as e:
                        st.error(f"Erreur lors de la mise à jour : {e}")

    
        except Exception as e:
            st.error(f"Erreur lors de la lecture du fichier : {e}")

    
    elif section == "ajout":
        st.subheader("Ajout d'une nouvelle fiche DRI")

        if st.session_state["ligne_temporaire"] is None:
            with st.form("formulaire"):
                date_reception = st.date_input("Date de réception (jj/mm/aaaa)", format="DD/MM/YYYY")
                reseau = st.text_input("Réseau")
                type_demande_input = st.radio("Type de demande", ["1 - DEPASSEMENT DE COUT", "2 - DEMANDE DE MA"])
                demande_type = "DEPASSEMENT DE COUT" if type_demande_input.startswith("1") else "DEMANDE DE MA"
                commentaire = st.text_area("Commentaire (optionnel)")
                fichier_kmz = st.file_uploader("Fichier BPE (KMZ ou archive ZIP de SHP)", type=["kmz", "zip"])
                fichier_gdb = st.file_uploader("Fichier GDB des parcelles réseau", type=["gdb", "zip"])
                fiche_dri_file = st.file_uploader("Fichier Excel de la fiche DRI", type="xlsx")
                fichier_bce = st.file_uploader("Fichier des entreprises (BCE) au format CSV UTF-8", type=["csv"])
                submit = st.form_submit_button("Ajouter la fiche")

            if submit and fiche_dri_file:
                try:
                    dri_wb = load_workbook(fiche_dri_file, data_only=True)
                    dri_ws = dri_wb.active

                    ligne = {
                        "DATE RECEPTION": date_reception.strftime("%d/%m/%Y"),
                        "RESEAU": reseau,
                        "RESPONSABLE PROD": dri_ws["C7"].value,
                        "COMMERCIAL": dri_ws["D16"].value,
                        "PROJET": dri_ws["D9"].value,
                        "TYPE DE DEMANDE": demande_type,
                        "COUT EXTENSION": dri_ws["D37"].value,
                        "COUT GLOBAL PROJET": dri_ws["D38"].value,
                        "OPERATEUR": dri_ws["D11"].value,
                        "GAIN DRI": dri_ws["G30"].value,
                        "ROI": round(dri_ws["G31"].value) if isinstance(dri_ws["G31"].value, (int, float)) else "ERREUR",
                        "NB CLIENTS AMORTISSEMENT": round((dri_ws["D38"].value - dri_ws["G30"].value) / 4000, 2)
                            if isinstance(dri_ws["D38"].value, (int, float)) and isinstance(dri_ws["G30"].value, (int, float)) else "ERREUR",
                        "COMMANDE": dri_ws["D10"].value,
                        "DATE TRAITEMENT": "",
                        "DELAI TRAITEMENT": "",
                        "ETAT GEOMARKETING": "",
                        "RESP GEOMARKET": "",
                        "CONCLUSION": "",
                        "COMMENTAIRE": commentaire,
                        "FICHIER_KMZ": fichier_kmz,
                        "FICHIER_GDB": fichier_gdb,
                        "FICHIER_BCE": fichier_bce,
                        "NOM_COMMANDE": nom_commande
                    }

                    st.session_state["ligne_temporaire"] = ligne

                except Exception as e:
                    st.error(f"Erreur lors de la lecture du fichier DRI : {e}")

        else:
            st.success("Fiche prête à être enregistrée :")
            df_temp = pd.DataFrame([st.session_state["ligne_temporaire"]])
            
            # Supprimer les colonnes qui posent problème (UploadedFile)
            colonnes_non_affichables = ["FICHIER_KMZ", "FICHIER_GDB"]
            df_temp_clean = df_temp.drop(columns=colonnes_non_affichables, errors="ignore")
            
            st.dataframe(df_temp_clean)

            # Traitement du fichier BPE (KMZ ou archive SHP) après soumission
            fichier_kmz = st.session_state["ligne_temporaire"].get("FICHIER_KMZ")
            nom_commande = st.session_state["ligne_temporaire"].get("NOM_COMMANDE")
            
            if fichier_kmz and nom_commande:
                if fichier_kmz.name.endswith(".zip"):
                    with tempfile.TemporaryDirectory() as tmpdir:
                        zip_path = os.path.join(tmpdir, "shapefile.zip")
                        with open(zip_path, "wb") as f:
                            f.write(fichier_kmz.read())
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            zip_ref.extractall(tmpdir)
            
                        shp_files = [f for f in os.listdir(tmpdir) if f.endswith(".shp")]
                        if not shp_files:
                            st.error("Aucun fichier .shp trouvé dans l'archive.")
                            gdf_kmz, chemin_shp = None, "Pas de .shp trouvé"
                        else:
                            shp_path = os.path.join(tmpdir, shp_files[0])
                            gdf_kmz, chemin_shp = traiter_bpe_shp(shp_path, nom_commande)
            
                elif fichier_kmz.name.endswith(".kmz"):
                    gdf_kmz, chemin_shp = traiter_kmz(fichier_kmz, nom_commande)
                else:
                    gdf_kmz, chemin_shp = None, "Format de fichier non supporté (utilisez KMZ ou ZIP de SHP)."
            
                if gdf_kmz is None:
                    st.error(f"Erreur dans le traitement du fichier BPE : {chemin_shp}")
                else:
                    st.success(f"Fichier BPE traité et exporté avec succès : {chemin_shp}")
                    st.session_state["shp_kmz_export"] = chemin_shp
            else:
                st.warning("Fichier BPE (KMZ ou ZIP) ou nom de commande manquant pour le traitement.")
                    

                
            fichier_gdb = st.session_state["ligne_temporaire"].get("FICHIER_GDB")
            nom_commande = st.session_state["ligne_temporaire"].get("NOM_COMMANDE")

            if fichier_gdb:
                thd_buffers, err_gdb = traiter_gdb_thd_zones(fichier_gdb)
                if thd_buffers is None:
                    st.error(f"Erreur GDB : {err_gdb}")
                else:
                    st.success("Zones THD extraites et bufferisées avec succès.")
                    st.session_state["thd_buffers"] = thd_buffers
            else:
                st.warning("Fichier GDB non fourni.")

            if gdf_kmz is not None and thd_buffers is not None:
                gdf_final, err_thd = attribuer_thd_extension(gdf_kmz, thd_buffers)
                if gdf_final is None:
                    st.error(f"Erreur attribution THD : {err_thd}")
                else:
                    st.success("Attribution THD effectuée avec succès.")
                    st.dataframe(gdf_final[["BPE", "COMMANDE", "DISTANCE", "THD_EXTENSION", "Z_BPE"]])
                    st.session_state["gdf_final"] = gdf_final  # Pour la synthèse finale

            # Vérification de l'existence de gdf_final avant de poursuivre
            if 'gdf_final' in locals() and gdf_final is not None:
                gdf_synth, shp_synth_path, err_synth = generer_synthese_finale(gdf_final, nom_commande)
                if gdf_synth is None:
                    st.error(f"Erreur synthèse finale : {err_synth}")
                else:
                    st.success("Fichier synthèse THD généré avec succès.")
                    st.dataframe(gdf_synth[["BPE", "THD_ACTUELLE", "THD_EXTENSION", "THD_FINALE", "GAIN_ENTREPRISE"]])
                    st.session_state["gdf_synthese"] = gdf_synth
                    st.session_state["shp_synthese"] = shp_synth_path
            
                    # Génération de l'image TCD MA
                    image_tcd_buf, err_img = generer_image_tcd_ma(gdf_synth, nom_commande)
                    if image_tcd_buf:
                        st.image(image_tcd_buf, caption="TCD MA généré")
                        st.session_state["img_tcd_ma"] = image_tcd_buf
                    else:
                        st.warning(f"TCD image non générée : {err_img}")
            
                    # Création du ZIP final
                    chemin_zip, err_zip = creer_zip_final(
                        st.session_state["shp_kmz_export"],
                        st.session_state["shp_synthese"],
                        st.session_state["img_tcd_ma"],
                        nom_commande
                    )
                    if chemin_zip:
                        with open(chemin_zip, "rb") as f:
                            st.download_button(
                                label="📦 Télécharger le ZIP final",
                                data=f,
                                file_name=f"{nom_commande}_EXPORT.zip",
                                mime="application/zip"
                            )
                    else:
                        st.error(f"Erreur lors de la création du ZIP : {err_zip}")
            
                    # Export SHP seul en option
                    with open(shp_synth_path, "rb") as f:
                        st.download_button(
                            label=f"Télécharger SHP Synthèse ({nom_commande})",
                            data=f,
                            file_name=f"{nom_commande}_synthese.shp",
                            mime="application/octet-stream"
                        )
            else:
                st.warning("Le traitement n'a pas abouti — gdf_final est introuvable ou vide.")


            col1, col2 = st.columns(2)
            with col1:
                if st.button("Enregistrer dans le fichier"):
                    try:
                        wb = load_workbook(suivi_file_path)
                        ws = wb.active
                        headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
                        new_row = [st.session_state["ligne_temporaire"].get(h, "") for h in headers]

                        last_row = max(
                            i for i, row in enumerate(ws.iter_rows(values_only=True), 1)
                            if any(cell is not None and str(cell).strip() != "" for cell in row)
                        )
                        next_row = last_row + 1

                        for col_idx, val in enumerate(new_row, start=1):
                            ws.cell(row=next_row, column=col_idx, value=val)

                        wb.save(suivi_file_path)
                        try_push_to_github()
                        st.success("Fiche enregistrée avec succès")
                        st.session_state["ligne_temporaire"] = None
                    except Exception as e:
                        st.error(f"Erreur lors de l'enregistrement : {e}")
            with col2:
                if st.button("Ajouter une autre fiche"):
                    st.session_state["ligne_temporaire"] = None
                    st.rerun()

    elif section == "analyse_donnees":
        st.subheader("Analyse des données - Synthèse Graphique et Statistique")

        try:
            import plotly.express as px
            df = pd.read_excel(suivi_file_path, engine="openpyxl")
            df.columns = [col.strip().upper() for col in df.columns]
            df["DATE RECEPTION"] = pd.to_datetime(df["DATE RECEPTION"], errors="coerce", dayfirst=True)
            df["ANNEE"] = df["DATE RECEPTION"].dt.year
            df["MOIS"] = df["DATE RECEPTION"].dt.strftime("%b")
            df["MOIS_NUM"] = df["DATE RECEPTION"].dt.month

            st.markdown("### 1. Coût global moyen & nombre de commandes par mois")
            grouped1 = df.groupby(["ANNEE", "MOIS", "MOIS_NUM"]).agg({
                "COUT GLOBAL PROJET": "mean",
                "COMMANDE": "count"
            }).reset_index().sort_values(by=["ANNEE", "MOIS_NUM"])
            grouped1 = grouped1.rename(columns={
                "COUT GLOBAL PROJET": "Coût moyen",
                "COMMANDE": "Nb commandes"
            })
            grouped1["Période"] = grouped1["ANNEE"].astype(str) + " - " + grouped1["MOIS"]
            st.dataframe(grouped1[["Période", "Coût moyen", "Nb commandes"]])


            st.markdown("### 2. Somme des coûts globaux par mois")
            grouped2 = df.groupby(["ANNEE", "MOIS", "MOIS_NUM"]).agg({
                "COUT GLOBAL PROJET": "sum",
                "COMMANDE": "count"
            }).reset_index().sort_values(by=["ANNEE", "MOIS_NUM"])
            grouped2 = grouped2.rename(columns={
                "COUT GLOBAL PROJET": "Coût total",
                "COMMANDE": "Nb commandes"
            })
            grouped2["Période"] = grouped2["ANNEE"].astype(str) + " - " + grouped2["MOIS"]
            st.dataframe(grouped2[["Période", "Coût total", "Nb commandes"]])


            st.markdown("### 3. Volume de commandes par responsable prod")
            commandes_prod = df["RESPONSABLE PROD"].value_counts().reset_index()
            commandes_prod.columns = ["Responsable", "Nb commandes"]
            st.dataframe(commandes_prod)
            st.bar_chart(commandes_prod.set_index("Responsable"))

            st.markdown("### 4. Volume de commandes par opérateur")
            commandes_operateur = df["OPERATEUR"].value_counts().reset_index()
            commandes_operateur.columns = ["Opérateur", "Nb commandes"]
            st.dataframe(commandes_operateur)
            st.bar_chart(commandes_operateur.set_index("Opérateur"))

            st.markdown("### 5. Répartition des commandes par date")
            repartition = df.groupby(["ANNEE", "MOIS", "MOIS_NUM"]).agg({"COMMANDE": "count"}).reset_index().sort_values(by=["ANNEE", "MOIS_NUM"])
            repartition.columns = ["ANNEE", "MOIS", "MOIS_NUM", "Nb commandes"]
            repartition["Période"] = repartition["ANNEE"].astype(str) + " - " + repartition["MOIS"]

            # Tableau
            st.dataframe(repartition[["Période", "Nb commandes"]])

            # Graphique
            fig = px.bar(repartition, x="Période", y="Nb commandes", title="Nb de commandes par mois")
            st.plotly_chart(fig, use_container_width=True)


            st.markdown("### 6. Délai de traitement moyen par mois")
            if "DELAI TRAITEMENT" in df.columns:
                df["DELAI TRAITEMENT"] = pd.to_numeric(df["DELAI TRAITEMENT"], errors="coerce")
                delai = df.groupby(["ANNEE", "MOIS", "MOIS_NUM"]).agg({"DELAI TRAITEMENT": "mean"}).reset_index().sort_values(by=["ANNEE", "MOIS_NUM"])
                delai["Période"] = delai["ANNEE"].astype(str) + " - " + delai["MOIS"]
                st.dataframe(delai[["Période", "DELAI TRAITEMENT"]].rename(columns={"DELAI TRAITEMENT": "Délai moyen"}))
                fig2 = px.line(delai, x="Période", y="DELAI TRAITEMENT", title="Délai de traitement moyen")
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("La colonne 'DELAI TRAITEMENT' est manquante dans le fichier.")

        except Exception as e:
            st.error(f"Erreur lors du chargement des analyses : {e}")

